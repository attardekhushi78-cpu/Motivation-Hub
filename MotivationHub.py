import os
import sys
import json
import random
import datetime
import requests
from pathlib import Path
from openpyxl import load_workbook, Workbook
import tkinter as tk
from tkinter import messagebox
from PIL import Image, ImageTk
import io

# Data file paths
DATA_FILES = {
    'advice': 'data/advice_data.xlsx',
    'books': 'data/books_data.xlsx',
    'quotes': 'data/quotes_data.xlsx',
    'dogs': 'data/dogs_data.xlsx'
}

# APIs
APIS = {
    'advice': 'https://api.adviceslip.com/advice',
    'books': 'https://potterapi-fedeperin.vercel.app/en/books',
    'quotes': 'https://hindi-quotes.vercel.app/random',
    'dogs': 'https://dog.ceo/api/breeds/image/random'
}

# Create required directories
def setup_directories():
    Path('data').mkdir(exist_ok=True)
    Path('logs').mkdir(exist_ok=True)
    Path('images').mkdir(exist_ok=True)

# Initialize Excel files with headers if they don’t exist
def init_excel_files():
    file_schemas = {
        'advice': ['ID', 'Advice', 'Timestamp'],
        'books': ['Title', 'Author', 'Description', 'Release Date', 'Pages', 'Timestamp'],
        'quotes': ['Quote', 'Author', 'Category', 'Timestamp'],
        'dogs': ['Image URL', 'Breed', 'Timestamp']
    }

    for file_type, headers in file_schemas.items():
        file_path = DATA_FILES[file_type]
        if not os.path.exists(file_path):
            wb = Workbook()
            ws = wb.active
            ws.title = f"{file_type.capitalize()} Data"
            ws.append(headers)
            wb.save(file_path)

# Get current time
def time_now():
    return datetime.datetime.now().strftime("%c")

# Log errors
def log_error(msg):
    with open("logs/app.log", "a") as f:
        f.write(f"[{time_now()}] ERROR: {msg}\n")

# API request (safe)
def make_api_request(api_name):
    try:
        response = requests.get(APIS[api_name], timeout=10)
        response.raise_for_status()

        return response.json()
    except Exception as e:
        log_error(f"API Error ({api_name}): {str(e)}")
        return None

# Save data to Excel
def save_to_excel(data_type, data):
    try:
        file_path = DATA_FILES[data_type]
        wb = load_workbook(file_path)
        ws = wb.active

        if data_type == 'advice':
            ws.append([data["slip"]["id"], data["slip"]["advice"], time_now()])

        elif data_type == 'books':
            book = random.choice(data)
            ws.append([
                book.get("title", "Unknown"),
                book.get("author", "Unknown"),
                book.get("description", "No description"),
                book.get("releaseDate", "Unknown"),
                book.get("pages", "Unknown"),
                time_now()
            ])

        elif data_type == 'quotes':
            ws.append([
                data.get("quote", "No quote"),
                data.get("author", "Unknown"),
                data.get("category", "General"),
                time_now()
            ])

        elif data_type == 'dogs':
            ws.append([
                data.get("message", "No URL"),
                "Unknown",
                time_now()
            ])

        wb.save(file_path)
        return True
    except Exception as e:
        log_error(f"Excel Save Error ({data_type}): {str(e)}")
        return False

#########################################
# GUI FUNCTIONS
#########################################

def get_advice():
    data = make_api_request("advice")
    if not data:
        messagebox.showerror("Error", "Could not fetch advice.")
        return

    slip = data["slip"]
    msg = f"Advice: {slip['advice']}\nID: {slip['id']}"
    messagebox.showinfo("Advice For Today", msg)
    save_to_excel("advice", data)


def get_book():
    data = make_api_request("books")
    if not data or not isinstance(data, list):
        messagebox.showerror("Error", "Could not fetch book info.")
        return

    book = random.choice(data)
    desc = book.get("description", "No description available.")

    msg = (
        f"Title: {book.get('title','Unknown')}\n"
        f"Author: {book.get('author', 'Unknown')}\n"
        f"Pages: {book.get('pages', 'Unknown')}\n"
        f"Release Date: {book.get('releaseDate', 'Unknown')}\n\n"
        f"{desc}"
    )

    messagebox.showinfo("Book Recommendation", msg)
    save_to_excel("books", data)


def get_quote():
    data = make_api_request("quotes")
    if not data:
        messagebox.showerror("Error", "Could not fetch quote.")
        return

    quote = f"“{data.get('quote','No quote')}”\n\n— {data.get('author','Unknown')}"
    messagebox.showinfo("Inspirational Quote", quote)
    save_to_excel("quotes", data)


def get_dog_image():
    data = make_api_request("dogs")
    if not data or data.get("status") != "success":
        messagebox.showerror("Error", "Could not fetch dog image.")
        return

    img_url = data["message"]
    filename = f"images/dog_{int(datetime.datetime.now().timestamp())}.jpg"

    try:
        img_bytes = requests.get(img_url, timeout=10).content
        with open(filename, "wb") as f:
            f.write(img_bytes)

        img = Image.open(io.BytesIO(img_bytes))
        img.thumbnail((300, 300))

        win = tk.Toplevel()
        win.title("Cute Dog Image")

        tk_img = ImageTk.PhotoImage(img)
        lbl = tk.Label(win, image=tk_img)
        lbl.image = tk_img
        lbl.pack()

        tk.Label(win, text=img_url).pack()

    except:
        messagebox.showerror("Error", "Image download failed.")
        log_error("Dog image download failed.")

    save_to_excel("dogs", data)


def get_everything():
    get_advice()
    get_quote()
    get_book()
    get_dog_image()
    messagebox.showinfo("All Done!", "🎉 Your Daily Motivation Pack is Ready! 🎉")


def show_statistics():
    stats = {}
    total = 0

    for name, path in DATA_FILES.items():
        if os.path.exists(path):
            wb = load_workbook(path)
            ws = wb.active
            count = ws.max_row - 1
            stats[name] = count
            total += count
        else:
            stats[name] = 0

    msg = (
        f"Total Entries: {total}\n"
        f"Advice: {stats['advice']}\n"
        f"Books: {stats['books']}\n"
        f"Quotes: {stats['quotes']}\n"
        f"Dogs: {stats['dogs']}"
    )

    messagebox.showinfo("Statistics", msg)

#########################################
# MAIN GUI WINDOW
#########################################

def main():
    setup_directories()
    init_excel_files()

    root = tk.Tk()
    root.title("Daily Motivation Hub")

    tk.Label(root, text="DAILY MOTIVATION HUB", font=("Arial", 18, "bold"), fg="green").pack(pady=15)

    tk.Button(root, text="Get Advice", width=30, command=get_advice).pack(pady=5)
    tk.Button(root, text="Get Quote", width=30, command=get_quote).pack(pady=5)
    tk.Button(root, text="Get Harry Potter Book", width=30, command=get_book).pack(pady=5)
    tk.Button(root, text="Get Cute Dog Image", width=30, command=get_dog_image).pack(pady=5)
    tk.Button(root, text="Get Everything", width=30, command=get_everything).pack(pady=5)
    tk.Button(root, text="Show Statistics", width=30, command=show_statistics).pack(pady=5)
    tk.Button(root, text="Exit", width=30, command=root.quit).pack(pady=10)

    root.mainloop()


if __name__ == "__main__":
    main()
